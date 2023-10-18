import datetime
import openpyxl
from telebot import formatting
import telebot
import os.path
from telebot import types

# Replace the placeholder data with your Atlas connection string. Be sure it includes
# a valid username and password! Note that in a production environment,
# you should not store your password in plain-text here.


from pymongo.mongo_client import MongoClient

uri = "mongodb+srv://elyorbek:elyorbek@cluster0.og6xz2d.mongodb.net/?retryWrites=true&w=majority"

# Create a new client and connect to the server
client = MongoClient(uri)

# Send a ping to confirm a successful connection
try:
    client.admin.command('ping')
    print("Pinged your deployment. You successfully connected to MongoDB!")
except Exception as e:
    print(e)

# use a database named "myDatabase"
db = client.myDatabase

bot_tolovi = 100000

# use a collection named "recipes"
my_collection = db["recipes"]

wb = openpyxl.Workbook()
# diskdan fayllarni yuklash

adminlar = "admin.xlsx"
ban_admin = "ban_admin.xlsx"
test_baza = "testlar.xlsx"

# openpyxl yordamida fayllarni ochish
admin_obj = openpyxl.load_workbook(adminlar)
ban_admin_obj = openpyxl.load_workbook(ban_admin)
testlar_wb_obj = openpyxl.load_workbook(test_baza)

# excel fayllardagi listlarni belgilash
admins_sheet = admin_obj.active
ban_admin_sheet = ban_admin_obj.active
testlar_sheet_obj = testlar_wb_obj.active

# telegram bot tokeni @BotFather
API_TOKEN = '1734744600:AAGIJiMZTmS2cof8Ee0mGqEelpvO9LH4_yE'
bot = telebot.TeleBot(API_TOKEN)

# statik ozgaruvchilar belgilandi
owner = 655955833  # bot egasi, super admin
admin_list = []
test_list = []  # test toplanadigan list

d = str(datetime.datetime.now())
dt = datetime.datetime.strptime(d, '%Y-%m-%d %H:%M:%S.%f')

# bazadan testlarni yuklash

m_row = testlar_sheet_obj.max_row
for i in range(1, m_row + 1):
    cell_obj = testlar_sheet_obj.cell(row=i, column=1)
    test_list.append(cell_obj.value)

test_kodi = 1000 + len(test_list)  # unikal kod

# bazadan adminlarni yuklash
admin_from_file = [[]]
m_row = admins_sheet.max_row
for i in range(1, m_row + 1):
    cell_obj = admins_sheet.cell(row=i, column=1)
    admin_list.append(cell_obj.value)


# admin hisobidan pul yechish
if datetime.datetime.now().day == 1:
    for admin in admin_list:
        myquery = {"_id": admin}
        mydoc = my_collection.find(myquery)
        for ad in mydoc:
            if ad['money'] >= bot_tolovi:
                my_collection.find_one_and_update({"_id": admin}, {
                    "$set": {"money": ad['money'] - bot_tolovi, "time": datetime.datetime.utcnow()}}, new=True)
                bot.send_message(chat_id=owner,
                                 text=f"{admin} ning balansidan 100 000 som yechib olindi.\noxirgi yangilanish {ad['time']}")
                bot.send_message(chat_id=admin,
                                 text=f"Balansingizdan 100 000 som yechib olindi.")
                myquery = {"_id": admin}
                mydoc = my_collection.find(myquery)
                for x in mydoc:
                    bot.send_message(chat_id=admin,
                                     text=f"Joriy balans: {x['money']}\noxirgi yangilanish {x['time']}")


            else:
                admin_list.remove(admin)
                bot.send_message(chat_id=owner,
                                 text=f"{admin} balansida pul yetarli emas, master darajasi bekor qilindi")
                bot.send_message(chat_id=admin,
                                 text=f"Balansingizda pul yetarli emas\nJoriy balans: {ad['money']} so'm\nMaster darajasi bekor qilindi\nBalansni to'ldirish uchun\n{formatting.format_text(formatting.hcode('9860160102726835'))} @ID1003\nraqamidan foydlaning\nDavom etish uchun /start ni boshing.",
                                 parse_mode='HTML')

        myquery = {"_id": admin}
        mydoc = my_collection.find(myquery)
        for x in mydoc:
            bot.send_message(chat_id=owner,
                             text=f"{admin} ning balansi: {x['money']}\noxirgi yangilanish {x['time']}")


@bot.message_handler(commands=['help', 'start'])
def send_welcome(message):
    if message.chat.id == owner:
        markup_owner = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup_owner.add('✳Admin belgilash', "✳Admin o'chirish", '✳Statistika', "✳Xisob")
        bot.send_message(message.chat.id,
                         f"Salom {message.from_user.first_name}. Sizni darajangiz admin.",
                         reply_markup=markup_owner)
    elif message.chat.id in admin_list:
        # bot.send_message(message.chat.id, "https://t.me/metamatem/39")
        # bot.send_video(chat_id=message.from_user.id, video=open("https://t.me/metamatem/39", 'rb'), supports_streaming=True)
        markup_admin = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup_admin.add('✳Test yaratish', '✳Fayl yuklang', '✳Statistika', "/testlar", '✳Balans')
        bot.send_message(message.chat.id,
                         "Sizni darajangiz Master",
                         reply_markup=markup_admin)
    else:
        # bot.send_message(message.chat.id, "https://t.me/metamatem/40")
        # bot.send_video(chat_id=message.from_user.id, video=open("https://t.me/metamatem/40", 'rb'), supports_streaming=True)
        markup_tinglovchi = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup_tinglovchi.add('✳Javobni tekshirish')
        bot.send_message(message.chat.id, text="Sizni darajangiz Tinglovchi", reply_markup=markup_tinglovchi)


@bot.message_handler(func=lambda message: True)
def all_messages(message):
    chat_id = message.chat.id
    try:
        if message.text == '':
            bot.copy_message(chat_id=-1001987872506, from_chat_id=message.chat.id, message_id=message.id)
        if chat_id == owner and message.text == "✳Admin belgilash":
            msg = bot.send_message(chat_id=message.chat.id, text="Admin telegram ID raqamini yuboring\n0123456789")
            bot.register_next_step_handler(msg, admin_belgilash)
        if chat_id == owner and message.text == "✳Admin o'chirish":
            msg = bot.send_message(chat_id=message.chat.id, text="Admin telegram ID raqamini yuboring\n0123456789")
            bot.register_next_step_handler(msg, admin_ochirish)
        if chat_id == owner and message.text == "✳Statistika":
            bot.send_message(chat_id=message.chat.id,
                             text=f'''{len(admin_list)}  ta admin\n{admin_list}\n{len(test_list)} ta test''')

        if chat_id == owner and message.text == "✳Xisob":
            msg = bot.send_message(chat_id=message.chat.id, text='Master ID raqamini yuboring')
            bot.register_next_step_handler(msg, xisob)

        if chat_id == owner and message.text == "+":
            msg = bot.send_message(chat_id=message.chat.id, text='Master ID raqami + summani yuboring')
            bot.register_next_step_handler(msg, xisob_qoshish)
        if chat_id in admin_list and message.text == "✳Test yaratish":
            msg = bot.send_message(chat_id=message.chat.id,
                                   text='''
👇👇👇 Yo'riqnoma.

1️⃣ Test yaratish uchun

Fan nomi*to'g'ri javoblar

ko`rinishida yuboring.

Misol:
Informatika*abbccdd...
yoki
Informatika*1a2d3c4a5b...

''')
            bot.register_next_step_handler(msg, test_yaratish)
        if chat_id in admin_list and message.text == "✳Statistika":
            admin_file = str(chat_id) + '.xlsx'
            my_file = open(admin_file, "rb")
            bot.send_document(chat_id=message.from_user.id, document=my_file,
                              caption=f"Oxirgi natijalar {dt + datetime.timedelta(hours=5)}\nMaster: {message.from_user.first_name}")

        if chat_id in admin_list and message.text == "✳Fayl yuklang":
            msg = bot.send_message(chat_id=message.chat.id, text="Test faylini yuboring")
            bot.register_next_step_handler(msg, bazaga_fayl_yuklash)

        if chat_id in admin_list and message.text == "/student":
            admin_list.remove(chat_id)

        if chat_id in admin_list and message.text == "✳Balans":
            my_query = {'_id': message.chat.id}
            mydoc = my_collection.find(my_query)
            if mydoc is not None:
                for d in mydoc:
                    bot.send_message(chat_id=message.chat.id,
                                     text=f"Balansingiz {d['money']} so'm\nBalansni to'ldirish uchun\n{formatting.format_text(formatting.hcode('9860160102726835'))} @ID1003\nraqamidan foydlaning",
                                     parse_mode='HTML')

        if message.text == "/master":
            new_master_id = message.chat.id
            if new_master_id not in admin_list:
                admin_belgilash(message)

            else:
                bot.send_message(chat_id=message.chat.id, text="Siz avvalroq master etib belgilangansiz.")

        if message.text == "✳Javobni tekshirish":
            msg = bot.send_message(chat_id=message.chat.id,
                                   text='''
👇👇👇 Yo'riqnoma.

1️⃣ Test javoblarini yuborish uchun

test kodi*abbccdd...
yoki
test kodi*1a2d3c4a5b...

kabi ko`rinishlarda yuboring.

Misol:
1234*abbccdd...
yoki
1234*1a2d3c4a5b...
''')
            bot.register_next_step_handler(msg, tekshirish)

    except Exception as e:
        bot.reply_to(message, str(e))


test_set = set(test_list)
test_list = list(test_set)

admin_set = set(admin_list)
admin_list = list(admin_set)


def bazaga_fayl_yuklash(message):
    bot.copy_message(chat_id=-1001987872506, from_chat_id=message.chat.id, message_id=message.id)
    bot.send_message(chat_id=message.chat.id, text="Fayl bazaga yuklandi. Uni /testlar buygug'i orqali topasiz")


def admin_belgilash(message):
    if message.chat.id not in admin_list:
        myquery = {"_id": message.chat.id}
        mydoc = my_collection.find(myquery)
        for k in mydoc:
            print(k['money'])
            if int(k['money']) > 0:
                print("k>0")
                try:
                    yangi_admin = int(message.chat.id)
                    admin_list.append(yangi_admin)

                    bazaga_yoziladi = [{'_id': int(message.chat.id), 'money': 0, 'time': datetime.datetime.utcnow()}]
                    result = my_collection.insert_many(bazaga_yoziladi)
                    inserted_count = len(result.inserted_ids)
                    # print("I inserted %x documents." % (inserted_count))

                    yangi_admin1 = [yangi_admin]
                    admins_sheet.append(yangi_admin1)
                    admin_obj.save(filename=adminlar)

                    admin_file_nomi = str(message.chat.id) + '.xlsx'
                    check_file = os.path.isfile(admin_file_nomi)
                    if not check_file:
                        wb.save(filename=admin_file_nomi)

                    bot.send_message(chat_id=owner,
                                     text=f"{yangi_admin} admin etib belgilandi, bazaga %x ta dokument qo'shildi" % (
                                         inserted_count))
                except Exception as e:
                    bot.reply_to(message, str(e))
            else:

                bot.send_message(chat_id=message.chat.id,
                                 text=f"Balansingizda pul yetarli emas\nJoriy balans: {k['money']} so'm\nMaster darajasi tiklanmadi\nBalansni to'ldirish uchun\n{formatting.format_text(formatting.hcode('9860160102726835'))} @ID1003\nraqamidan foydlaning.",
                                 parse_mode='HTML')
    else:
        bot.reply_to(message, f"{message.chat.id} avvalroq master etib belgilangan")


def xisob(message):
    myquery = {"_id": int(message.text)}
    mydoc = my_collection.find(myquery)
    for x in mydoc:
        print(x)
        print(x['money'])
        bot.send_message(chat_id=message.chat.id,
                         text=f"{message.text} ning balansi: {x['money']}\noxirgi yangilanish {x['time']}")


def xisob_qoshish(message):
    text = str(message.text).replace(' ', '')
    master = int(text[:text.find('+')])
    summa = int(text[text.find('+') + 1:])
    myquery1 = {"_id": master}
    mydoc1 = my_collection.find(myquery1)
    for xl in mydoc1:
        my_collection.find_one_and_update({"_id": master},
                                          {"$set": {"money": xl['money'] + summa, "time": datetime.datetime.utcnow()}},
                                          new=True)
        bot.send_message(chat_id=message.chat.id, text=f"{master} balansi {summa} ga oshirildi")
        # bot.send_message(chat_id=master, text=f"Balansingiz {summa} ga oshirildi")
        if master not in admin_list:
            myquery1 = {"_id": master}
            mydoc1 = my_collection.find(myquery1)
            for xs in mydoc1:
                if int(xs['money']) >= bot_tolovi:
                    my_collection.find_one_and_update({"_id": master}, {
                        "$set": {"money": xs['money'] - bot_tolovi, "time": datetime.datetime.utcnow()}}, new=True)
                    admin_list.append(master)
                    bot.send_message(chat_id=message.chat.id, text=f"{master} ning master darajasi tiklandi")
                    bot.send_message(chat_id=master,
                                     text=f"Hisobingiz {summa} ga to'ldirildi\nO'tgan foydalanilgan davr uchun {0 - xl['money']} so'm ushlab qolindi.")
                    bot.send_message(chat_id=master, text=f"Joriy balans: {xs['money']}")
                    bot.send_message(chat_id=master,
                                     text=f"Master darajasi tiklandi\nDavom etish uchun /start ni bosing.")

                else:
                    bot.send_message(chat_id=master,
                                     text=f"Hisobingiz {summa} ga to'ldirildi\nJoriy balans {xs['money']} so'm\nMaster darajasi tiklanmadi.")

        else:
            mydoc1 = my_collection.find(myquery1)
            for s in mydoc1:
                bot.send_message(chat_id=message.chat.id,
                                 text=f"{master} ning balansi: {s['money']}\noxirgi yangilanish {s['time']}")
                bot.send_message(chat_id=master,
                                 text=f"Hisobingiz {summa} ga to'ldirildi\nJoriy balans: {s['money']} so'm")


def admin_ochirish(message):
    try:
        eski_admin = int(message.text)
        for uchadigan_admin in admin_list:
            if uchadigan_admin == eski_admin:
                admin_list.remove(uchadigan_admin)
        eski_admin1 = [eski_admin]
        ban_admin_sheet.append(eski_admin1)
        ban_admin_obj.save(filename=ban_admin)
        bot.send_message(chat_id=owner, text=f"{eski_admin} adminlikdan o'chirildi")
    except Exception as e:
        bot.reply_to(message, str(e))


def test_yaratish(message):
    global test_kodi, test_list
    try:
        text1 = str(message.text).lower()
        text2 = text1.replace(' ', '')
        text3 = text2.replace('а', 'a')
        text4 = text3.replace('б', 'b')
        text5 = text4.replace('с', 'c')
        text = text5.replace('д', 'd')
        if "*" in text:
            fan = text[:text.find('*')]
            javoblar = text[text.find("*") + 1:]
            modified_string = ''.join([im for im in javoblar if not im.isdigit()])
            modified_string1 = modified_string.replace(' ', '')
            test_kodi += 1
            namuna = str(fan) + '*' + str(message.chat.id) + "*" + str(test_kodi) + "*" + modified_string1
            test_list.append(namuna)
            test_set = set(test_list)
            test_list = list(test_set)
            namuna1 = [namuna]
            testlar_sheet_obj.append(namuna1)
            testlar_wb_obj.save(filename=test_baza)
            bot.send_message(chat_id=message.chat.id,
                             text=f'''✅Test bazaga qoshildi\n\nTest kodi: {test_kodi}\nSavollar soni: {len(modified_string)} ta\n\nTestda qatnashuvchilar quyidagi ko`rinishda javob yuborishlari mumkin:\n\n{test_kodi}*aabbcc... ({len(modified_string)} ta)\n\nyoki\n{test_kodi}*1a2a3b4b5c6c...  \n\nTest kodini biriktirib, faylini ham yuboring.''')

        else:
            bot.send_message(chat_id=message.chat.id, text="test yaratish xatoligi")
    except Exception as e:
        bot.reply_to(message, text="test yaratish xatoligi")


def tekshirish(message):
    tekshir = ""
    master = ''
    test_fani = ''
    n = 1
    b = 0
    try:
        text1 = str(message.text).lower()
        text2 = text1.replace(' ', '')
        text3 = text2.replace('а', 'a')
        text4 = text3.replace('б', 'b')
        text5 = text4.replace('с', 'c')
        text = text5.replace('д', 'd')
        if "*" in text:
            javoblar = text[text.find("*") + 1:]
            modified_string = ''.join([mi for mi in javoblar if not mi.isdigit()])
            modified_string1 = modified_string.replace(' ', '')
            mytest_kodi = text[:text.find("*")]
            mynamuna = str(mytest_kodi) + "*" + modified_string1
            for sd in test_list:
                ef = str(sd)
                fani = ef[:ef.find('*')]

                togri_javob1 = ef.replace(fani + '*', '')
                master_id = togri_javob1[:togri_javob1.find("*")]

                togri_javob = togri_javob1.replace(master_id + '*', '')
                if mytest_kodi == togri_javob[:togri_javob.find("*")] and len(mynamuna) == len(togri_javob):
                    master = master_id
                    test_fani = fani

                    for il, j in zip(mynamuna[mynamuna.find("*") + 1:], togri_javob[togri_javob.find("*") + 1:]):
                        if il == j:
                            tekshir += ("{} {} ✅\n".format(n, il))  # ✅⛔️
                            b += 1
                        else:
                            tekshir += ("{} {} ⛔\n".format(n, il))
                        n += 1
            tinglovchi_id = [message.chat.id]
            ishlash_vaqti = [str(dt + datetime.timedelta(hours=5))]
            ism = [message.from_user.first_name]
            fan = [test_fani]
            test_kodi1 = [mytest_kodi]
            javoblar = [tekshir]
            ball = [str(b)]
            test_natija_fayli = tinglovchi_id + ishlash_vaqti + ism + fan + test_kodi1 + javoblar + ball
            wb_nomi = str(master) + '.xlsx'
            yoziladigan_file = openpyxl.load_workbook(wb_nomi)
            yoziladigan_list = yoziladigan_file.active
            yoziladigan_list.append(test_natija_fayli)
            yoziladigan_file.save(filename=wb_nomi)
            bot.send_message(chat_id=message.chat.id, text=f"{tekshir}\n{b} ball to'pladingiz.")
        else:
            bot.send_message(chat_id=message.chat.id, text="test tekshirish xatoligi")
    except Exception as e:
        bot.reply_to(message, text="test tekshirish xatoligi")

bot.infinity_polling()